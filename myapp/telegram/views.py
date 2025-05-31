# Django imports
from django.shortcuts import render
from django.http import HttpResponse
from django.http import JsonResponse
from django.core.paginator import Paginator
from asgiref.sync import sync_to_async

# Django users imports
from django.contrib.auth.decorators import login_required

# Forms
from myapp.telegram.forms import TelegramForm

# Models
from .models import TelegramChannel, TelegramPost

# Filters
#from .filters import TelegramFilter

# Regular expressions
import re

# Date and time
from datetime import datetime
from datetime import timezone
import time

# Telegram related imports
from telethon import TelegramClient
from telethon import utils
from telethon import types
from telethon import functions

# Logging
import logging

# Excel processing
import openpyxl

# JSON handling
import json

# UUID generation
import uuid

# Operating system utilities
import os

# Environment variables
from dotenv import load_dotenv

# Machine learning imports
from myapp.telegram.machine_learning.train_model import load_model
from myapp.telegram.machine_learning.train_model import predict_category
from myapp.telegram.machine_learning.train_model import update_model
from myapp.telegram.machine_learning.train_model import get_unique_categories
from myapp.telegram.machine_learning.train_model import export_model

# Image processing
from PIL import Image
import io
from collections import defaultdict
from statistics import mean

from asgiref.sync import sync_to_async

# Views imports
from myapp.views import cleanup_temp_data

load_dotenv()
api_id = os.getenv("API_ID")  
api_hash = os.getenv("API_HASH")  

# Настройка логирования
logger = logging.getLogger(__name__)

# Загрузка модели и векторизатора при старте
model, vectorizer = load_model()

def extract_username(url):
    match = re.search(r'https://t.me/([^/?]+)', url)
    if match:
        return match.group(1)
    raise ValueError("Invalid Telegram URL")


async def fetch_telegram_data(channel_url, start_date=None, end_date=None):
    global model, vectorizer
    client = TelegramClient('session_name', api_id, api_hash)
    try:
        if not client.is_connected():
            await client.connect()
        if not await client.is_user_authorized():
            logger.warning("Клиент Telegram не авторизован. Пожалуйста, авторизуйтесь вручную.")
            print("Введите номер телефона (например, +79991234567): ")
            phone = input().strip()
            await client.start(phone=phone)

        if start_date:
            start_date = datetime.combine(start_date, datetime.min.time()).replace(tzinfo=timezone.utc)
            logger.debug(f"Дата начала установлена на: {start_date}")
        if end_date:
            end_date = datetime.combine(end_date, datetime.max.time()).replace(tzinfo=timezone.utc)
            logger.debug(f"Дата окончания установлена на: {end_date}")

        if "https://t.me/" in channel_url:
            username = channel_url.split('/')[-1]
        elif "https://web.telegram.org/k/#" in channel_url:
            channel_id = int(channel_url.split('#')[-1])
            username = utils.get_peer_id(types.PeerChannel(channel_id))
        else:
            username = extract_username(channel_url)

        logger.debug(f"Получение сущности для username: {username}")
        entity = await client.get_entity(username)
        data = []
        combined_message = None

        # Получение количества подписчиков
        subscriber_count = 0
        try:
            full_channel = await client(functions.channels.GetFullChannelRequest(channel=entity))
            subscriber_count = int(getattr(full_channel.full_chat, 'participants_count', 0))
            logger.debug(f"Количество подписчиков из GetFullChannelRequest для {entity.title}: {subscriber_count}")
        except Exception as e:
            logger.error(f"Ошибка получения количества подписчиков для канала {entity.id}: {str(e)}")
            subscriber_count = 0

        if subscriber_count == 0:
            logger.warning(f"Не удалось получить валидное количество подписчиков для канала {entity.title}. Метрики будут равны 0. Убедитесь, что клиент имеет разрешение 'Просмотр статистики канала'.")

        # Загрузка аватара канала
        avatar_path = None
        avatar_dir = os.path.join('temp_data', 'avatars')
        os.makedirs(avatar_dir, exist_ok=True)
        avatar_filename = f"channel_{entity.id}.jpg"
        avatar_full_path = os.path.join(avatar_dir, avatar_filename)

        if hasattr(entity, 'photo') and entity.photo:
            logger.debug(f"Загрузка фото профиля для канала {entity.id}")
            try:
                photo = await client.download_profile_photo(entity, file=avatar_full_path)
                if photo and os.path.exists(avatar_full_path):
                    logger.debug(f"Фото профиля загружено в {avatar_full_path}")
                    try:
                        with Image.open(avatar_full_path) as img:
                            img = img.resize((50, 50), Image.LANCZOS)
                            img.save(avatar_full_path, 'JPEG', quality=85)
                        avatar_path = f"avatars/{avatar_filename}"
                        logger.info(f"Аватар обработан и сохранён для канала {entity.id}: {avatar_path}")
                    except Exception as e:
                        logger.error(f"Ошибка обработки аватара для канала {entity.id}: {str(e)}")
                        avatar_path = None
                else:
                    logger.warning(f"Аватар не загружен для канала {entity.id} (photo={photo})")
                    avatar_path = None
            except Exception as e:
                logger.error(f"Ошибка загрузки аватара для канала {entity.id}: {str(e)}")
                avatar_path = None
        else:
            logger.debug(f"Фото профиля недоступно для канала {entity.id}")
            avatar_path = None

        total_engagement = 0
        total_comments = 0
        post_count = 0

        async for post in client.iter_messages(entity, reverse=True, offset_date=start_date):
            if end_date and post.date and post.date > end_date:
                logger.debug(f"Пропуск поста {post.id} так как дата {post.date} позже end_date {end_date}")
                break

            # Определение типа поста
            post_type = 'text'
            if post.media:
                if isinstance(post.media, types.MessageMediaPhoto):
                    post_type = 'image'
                elif isinstance(post.media, types.MessageMediaDocument) and post.media.document.mime_type.startswith('video'):
                    post_type = 'video'

            message_text = post.message if post.message else 'N/A'
            # Игнорируем посты, которые являются только медиа без текста или вложенными медиа в группе
            if message_text == 'N/A' and (not post.grouped_id or (combined_message and combined_message.get('post_id') != post.id)):
                logger.debug(f"Игнорирование поста {post.id}: только медиа без текста или вложенный файл (тип: {post_type})")
                if post.grouped_id and combined_message:
                    combined_message['message'] += f"\n[{post_type.capitalize()}: {post.id}]"
                    combined_message['link'] = f"https://t.me/{entity.username}/{post.id}"
                continue

            if post.media and isinstance(post.media, types.MessageMediaPhoto) and not post.message and combined_message and combined_message.get('message') != 'N/A':
                combined_message['message'] += f"\n[Изображение: {post.media.photo.id}]"
                combined_message['link'] = f"https://t.me/{entity.username}/{post.id}"
                continue

            comments_count = post.replies.replies if hasattr(post, 'replies') and post.replies else 0
            reactions_count = sum(reaction.count for reaction in post.reactions.results) if hasattr(post, 'reactions') and post.reactions else 0
            forwards_count = post.forwards if hasattr(post, 'forwards') else 0
            views = post.views if hasattr(post, 'views') else 0

            logger.debug(f"Пост {post.id}: comments={comments_count}, reactions={reactions_count}, forwards={forwards_count}, views={views}, type={post_type}")

            category = predict_category(message_text, model, vectorizer) if message_text != 'N/A' else 'N/A'

            # Расчёт метрик
            post_engagement = reactions_count + forwards_count + comments_count
            total_engagement += post_engagement
            total_comments += comments_count
            post_count += 1

            er_post = (post_engagement / subscriber_count * 100) if subscriber_count > 0 and post_engagement > 0 else 0
            er_view = (post_engagement / views * 100) if views > 0 and post_engagement > 0 else 0
            vr_post = (views / subscriber_count * 100) if subscriber_count > 0 and views > 0 else 0
            logger.debug(f"Пост {post.id}: engagement={post_engagement}, er_post={er_post}%, er_view={er_view}%, vr_post={vr_post}%")

            combined_message = {
                'title': entity.title,
                'id': entity.id,
                'post_id': post.id,
                'date': post.date.strftime('%Y-%m-%d %H:%M:%S') if post.date else 'N/A',
                'message': message_text,
                'link': f"https://t.me/{entity.username}/{post.id}" if entity.username else f"https://t.me/c/{entity.id}/{post.id}",
                'views': views,
                'reactions': reactions_count,
                'forwards': forwards_count,
                'comments_count': comments_count,
                'category': category,
                'avatar': avatar_path,
                'er_post': round(er_post, 2),
                'er_view': round(er_view, 2),
                'vr_post': round(vr_post, 2),
                'type': post_type,
                'subscribers': subscriber_count
            }

            data.append(combined_message)

        # Расчёт TR
        tr = (total_comments / subscriber_count / post_count * 100) if post_count > 0 and subscriber_count > 0 else 0
        logger.debug(f"Расчёт TR: total_comments={total_comments}, subscriber_count={subscriber_count}, post_count={post_count}, tr={tr}%")
        for item in data:
            item['tr'] = round(tr, 2)

        data_id = str(uuid.uuid4())
        os.makedirs('temp_data', exist_ok=True)
        with open(f'temp_data/{data_id}.json', 'w', encoding='utf-8') as f:
            json.dump(data, f)

        logger.debug(f"Получены данные для {len(data)} постов из канала {entity.title}")
        logger.debug(f"ID данных: {data_id}")
        logger.debug(f"Пути аватаров в данных: {[item['avatar'] for item in data if item['avatar']]}")

        return data, data_id
    except Exception as e:
        logger.error(f"Ошибка получения данных для канала {channel_url}: {str(e)}")
        raise
    finally:
        if client.is_connected():
            await client.disconnect()


@login_required
async def telegram_view(request):
    cleanup_temp_data()
    
    global model, vectorizer
    print(f"Model in telegram_view: {model}, Vectorizer: {vectorizer}")
    if model is None or vectorizer is None:
        logger.error("Model or vectorizer is None. Attempting to reload...")
        model, vectorizer = load_model()

    filters = request.GET.getlist('filter')
    category_filters = request.GET.getlist('category_filter')
    channel_url = request.GET.get('channel_url', '')
    start_date = request.GET.get('start_date', '')
    end_date = request.GET.get('end_date', '')
    data_id = request.GET.get('data_id', '')
    sort_by = request.GET.get('sort_by', 'date')
    sort_direction = request.GET.get('sort_direction', 'desc')

    if request.method == 'POST':
        form = TelegramForm(request.POST)
        if form.is_valid():
            channel_urls = form.cleaned_data['channel_url'].strip().split('\n')
            start_date = form.cleaned_data['start_date']
            end_date = form.cleaned_data['end_date']
            all_data = []
            data_id_list = []

            for channel_url in channel_urls:
                channel_url = channel_url.strip()
                if channel_url:
                    data, data_id = await fetch_telegram_data(channel_url, start_date, end_date)
                    all_data.extend(data)
                    data_id_list.append(data_id)
            
            combined_data_id = str(uuid.uuid4())
            with open(f'temp_data/{combined_data_id}.json', 'w', encoding='utf-8') as f:
                json.dump(all_data, f)

            unique_titles = sorted(list(set(item['title'] for item in all_data)))
            logger.debug(f"Combined data passed to template: {len(all_data)} posts")
            logger.debug(f"Combined Data ID passed to template: {combined_data_id}")
            
            filtered_data = [
                item for item in all_data 
                if (not filters or 'all' in filters or item['title'] in filters) and
                   (not category_filters or 'all' in category_filters or item['category'] in category_filters)
            ]
            
            reverse = sort_direction == 'desc'
            if sort_by == 'channel':
                filtered_data.sort(key=lambda x: x['title'], reverse=reverse)
            elif sort_by == 'postid':
                filtered_data.sort(key=lambda x: x['post_id'], reverse=reverse)
            elif sort_by == 'date':
                filtered_data.sort(key=lambda x: x['date'], reverse=reverse)
            elif sort_by == 'category':
                filtered_data.sort(key=lambda x: x['category'], reverse=reverse)

            unique_categories_in_data = sorted(list(set(item['category'] for item in filtered_data if item['category'])))
            paginator = Paginator(filtered_data, 20)
            page_number = request.GET.get('page')
            page_obj = paginator.get_page(page_number)

            # Подготовка данных для графиков
            days_map = {
                'Понедельник': 'Monday',
                'Вторник': 'Tuesday',
                'Среда': 'Wednesday',
                'Четверг': 'Thursday',
                'Пятница': 'Friday',
                'Суббота': 'Saturday',
                'Воскресенье': 'Sunday'
            }
            publications_by_day = defaultdict(int)
            er_by_day = defaultdict(list)
            vr_by_day = defaultdict(list)
            content_types = defaultdict(int)
            top_posts_data = []

            for item in all_data:
                date = datetime.strptime(item['date'], '%Y-%m-%d %H:%M:%S')
                day_name = date.strftime('%A')
                for ru_day, en_day in days_map.items():
                    if en_day == day_name:
                        publications_by_day[ru_day] += 1
                        try:
                            er_by_day[ru_day].append(float(item.get('er_post', 0)))
                        except (ValueError, TypeError):
                            er_by_day[ru_day].append(0)
                        try:
                            vr_by_day[ru_day].append(float(item.get('vr_post', 0)))
                        except (ValueError, TypeError):
                            vr_by_day[ru_day].append(0)
                content_type = item.get('type', 'unknown').lower()
                if content_type in ['image', 'video', 'text']:
                    content_types[content_type] += 1

            er_by_day_data = {day: mean(ers) for day, ers in er_by_day.items() if ers}
            vr_by_day_data = {day: mean(vrs) for day, vrs in vr_by_day.items() if vrs}
            content_types_data = {content_type: (count / len(all_data)) * 100 for content_type, count in content_types.items() if count > 0}

            posts_by_channel_day = defaultdict(list)
            for item in all_data:
                date = datetime.strptime(item['date'], '%Y-%m-%d %H:%M:%S').strftime('%Y-%m-%d')
                channel = item['title']
                key = (date, channel)
                posts_by_channel_day[key].append({
                    'post_id': item['post_id'],
                    'message': item['message'],
                    'vr_post': float(item['vr_post'])
                })
            
            for (date, channel), posts in posts_by_channel_day.items():
                top_posts = sorted(posts, key=lambda x: x['vr_post'], reverse=True)[:3]
                for post in top_posts:
                    top_posts_data.append({
                        'date': date,
                        'channel': channel,
                        'post_id': post['post_id'],
                        'message': post['message'],
                        'vr_post': post['vr_post']
                    })

            for day in days_map.keys():
                if day not in publications_by_day:
                    publications_by_day[day] = 0
                if day not in er_by_day_data:
                    er_by_day_data[day] = 0
                if day not in vr_by_day_data:
                    vr_by_day_data[day] = 0

            context = {
                'data_id': data_id,
                'form': form,
                'parsed_data': page_obj,
                'data_id': combined_data_id,
                'filters': filters,
                'category_filters': category_filters,
                'unique_titles': unique_titles,
                'channel_url': request.POST.get('channel_url'),
                'start_date': request.POST.get('start_date'),
                'end_date': request.POST.get('end_date'),
                'sort_by': sort_by,
                'sort_direction': sort_direction,
                'unique_categories': get_unique_categories(),
                'unique_categories_in_data': unique_categories_in_data,
                'publications_by_day': json.dumps(dict(publications_by_day)),
                'er_by_day': json.dumps(er_by_day_data),
                'vr_by_day': json.dumps(vr_by_day_data),
                'content_types': json.dumps(content_types_data),
                'top_posts': top_posts_data,
            }
            print("Rendering telegram_main.html with data_id:", data_id)  # Отладка
            # Оборачиваем render в sync_to_async
            return await sync_to_async(render)(request, 'myapp/telegram/telegram_main.html', context)
    else:
        form = TelegramForm(initial={
            'channel_url': channel_url,
            'start_date': start_date,
            'end_date': end_date,
        })
    
    parsed_data = []
    if data_id:
        try:
            with open(f'temp_data/{data_id}.json', 'r', encoding='utf-8') as f:
                parsed_data = json.load(f)
            logger.debug(f"Loaded parsed data with {len(parsed_data)} posts")
        except FileNotFoundError:
            logger.error(f"Data file not found: temp_data/{data_id}.json")
    
    unique_titles = sorted(list(set(item['title'] for item in parsed_data)))
    
    filtered_data = [
        item for item in parsed_data 
        if (not filters or 'all' in filters or item['title'] in filters) and
           (not category_filters or 'all' in category_filters or item['category'] in category_filters)
    ]
    
    reverse = sort_direction == 'desc'
    if sort_by == 'channel':
        filtered_data.sort(key=lambda x: x['title'], reverse=reverse)
    elif sort_by == 'postid':
        filtered_data.sort(key=lambda x: x['post_id'], reverse=reverse)
    elif sort_by == 'date':
        filtered_data.sort(key=lambda x: x['date'], reverse=reverse)
    elif sort_by == 'category':
        filtered_data.sort(key=lambda x: x['category'], reverse=reverse)

    unique_categories_in_data = sorted(list(set(item['category'] for item in filtered_data if item['category'])))
    paginator = Paginator(filtered_data, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    # Подготовка данных для графиков
    days_map = {
        'Понедельник': 'Monday',
        'Вторник': 'Tuesday',
        'Среда': 'Wednesday',
        'Четверг': 'Thursday',
        'Пятница': 'Friday',
        'Суббота': 'Saturday',
        'Воскресенье': 'Sunday'
    }
    publications_by_day = defaultdict(int)
    er_by_day = defaultdict(list)
    vr_by_day = defaultdict(list)
    content_types = defaultdict(int)
    top_posts_data = []

    for item in parsed_data:
        date = datetime.strptime(item['date'], '%Y-%m-%d %H:%M:%S')
        day_name = date.strftime('%A')
        for ru_day, en_day in days_map.items():
            if en_day == day_name:
                publications_by_day[ru_day] += 1
                try:
                    er_by_day[ru_day].append(float(item.get('er_post', 0)))
                except (ValueError, TypeError):
                    er_by_day[ru_day].append(0)
                try:
                    vr_by_day[ru_day].append(float(item.get('vr_post', 0)))
                except (ValueError, TypeError):
                    vr_by_day[ru_day].append(0)
        content_type = item.get('type', 'unknown').lower()
        if content_type in ['image', 'video', 'text']:
            content_types[content_type] += 1

    er_by_day_data = {day: mean(ers) for day, ers in er_by_day.items() if ers}
    vr_by_day_data = {day: mean(vrs) for day, vrs in vr_by_day.items() if vrs}
    content_types_data = {content_type: (count / len(parsed_data)) * 100 for content_type, count in content_types.items() if count > 0}

    posts_by_channel_day = defaultdict(list)
    for item in parsed_data:
        date = datetime.strptime(item['date'], '%Y-%m-%d %H:%M:%S').strftime('%Y-%m-%d')
        channel = item['title']
        key = (date, channel)
        posts_by_channel_day[key].append({
            'post_id': item['post_id'],
            'message': item['message'],
            'vr_post': float(item['vr_post'])
        })
    
    for (date, channel), posts in posts_by_channel_day.items():
        top_posts = sorted(posts, key=lambda x: x['vr_post'], reverse=True)[:3]
        for post in top_posts:
            top_posts_data.append({
                'date': date,
                'channel': channel,
                'post_id': post['post_id'],
                'message': post['message'],
                'vr_post': post['vr_post']
            })

    for day in days_map.keys():
        if day not in publications_by_day:
            publications_by_day[day] = 0
        if day not in er_by_day_data:
            er_by_day_data[day] = 0
        if day not in vr_by_day_data:
            vr_by_day_data[day] = 0

    context = {
        'form': form,
        'parsed_data': page_obj,
        'data_id': data_id,
        'filters': filters,
        'category_filters': category_filters,
        'unique_titles': unique_titles,
        'channel_url': channel_url,
        'start_date': start_date,
        'end_date': end_date,
        'sort_by': sort_by,
        'sort_direction': sort_direction,
        'unique_categories': get_unique_categories(),
        'unique_categories_in_data': unique_categories_in_data,
        'publications_by_day': json.dumps(dict(publications_by_day)),
        'er_by_day': json.dumps(er_by_day_data),
        'vr_by_day': json.dumps(vr_by_day_data),
        'content_types': json.dumps(content_types_data),
        'top_posts': top_posts_data,
    }
    # Оборачиваем render в sync_to_async
    return await sync_to_async(render)(request, 'myapp/telegram/telegram_main.html', context)


def export_to_excel(request):
    data_id = request.GET.get('data_id')
    if not data_id:
        return HttpResponse("No data ID provided.", status=400)

    file_path = f"temp_data/{data_id}.json"
    try:
        with open(file_path, 'r', encoding='utf-8') as f:
            parsed_data = json.load(f)
    except FileNotFoundError:
        return HttpResponse("Data file not found.", status=404)

    workbook = openpyxl.Workbook()
    
    # Get date range for filename
    dates = [datetime.strptime(item['date'], '%Y-%m-%d %H:%M:%S') for item in parsed_data]
    min_date = min(dates).strftime('%Y%m%d')
    max_date = max(dates).strftime('%Y%m%d')

    # Sheet 1: Original Data
    sheet1 = workbook.active
    sheet1.title = 'Исходные данные'
    headers = ['Канал', 'ID Канала', 'ID Поста', 'Дата', 'Сообщение', 'Категория', 'Ссылка', 'Просмотров', 'Реакции', 'Пересылки', 'Комментарии']
    sheet1.append(headers)

    # 2) Расчет процента всех категорий на основе parsed_data
    category_counts = defaultdict(int)
    total_posts = len(parsed_data)
    for item in parsed_data:
        category = item.get('category', 'N/A')
        category_counts[category] += 1
    category_percentages = {cat: (count / total_posts * 100) for cat, count in category_counts.items()}

    for item in parsed_data:
        row = [
            item['title'], 
            item['id'], 
            item['post_id'], 
            item['date'], 
            item['message'], 
            item['category'],
            item['link'], 
            item['views'], 
            item['reactions'], 
            item['forwards'],
            item['comments_count']
        ]
        sheet1.append(row)

    # Добавляем строку с процентами категорий в конец Sheet 1
    sheet1.append([''] * (len(headers) - 1) + ['Процент категорий'])
    for cat, percent in category_percentages.items():
        sheet1.append([''] * (len(headers) - 1) + [f"{cat}: {percent:.2f}%"])

    # Sheet 2: Publications by Day of Week
    sheet2 = workbook.create_sheet('Количество публикаций')
    sheet2.append(['День недели', 'Количество публикаций'])
    publications_by_day = defaultdict(int)
    days_map = {
        'Понедельник': 'Monday',
        'Вторник': 'Tuesday',
        'Среда': 'Wednesday',
        'Четверг': 'Thursday',
        'Пятница': 'Friday',
        'Суббота': 'Saturday',
        'Воскресенье': 'Sunday'
    }
    for item in parsed_data:
        date = datetime.strptime(item['date'], '%Y-%m-%d %H:%M:%S')
        day_name = date.strftime('%A')
        for ru_day, en_day in days_map.items():
            if en_day == day_name:
                publications_by_day[ru_day] += 1
    for day, count in publications_by_day.items():
        sheet2.append([day, count])

    # Sheet 3: Average ER Post by Day
    sheet3 = workbook.create_sheet('Ср. ERpost по дню')
    sheet3.append(['День недели', 'Средний ER Post'])
    er_by_day = defaultdict(list)
    for item in parsed_data:
        date = datetime.strptime(item['date'], '%Y-%m-%d %H:%M:%S')
        day_name = date.strftime('%A')
        for ru_day, en_day in days_map.items():
            if en_day == day_name:
                er_by_day[ru_day].append(float(item['er_post']))
    for day, ers in er_by_day.items():
        sheet3.append([day, mean(ers)])

    # Sheet 4: Average VR Post by Day
    sheet4 = workbook.create_sheet('Ср. VRpost по дню')
    sheet4.append(['День недели', 'Средний VR Post'])
    vr_by_day = defaultdict(list)
    for item in parsed_data:
        date = datetime.strptime(item['date'], '%Y-%m-%d %H:%M:%S')
        day_name = date.strftime('%A')
        for ru_day, en_day in days_map.items():
            if en_day == day_name:
                vr_by_day[ru_day].append(float(item['vr_post']))
    for day, vrs in vr_by_day.items():
        sheet4.append([day, mean(vrs)])

    # Sheet 5: Content Type Distribution
    sheet5 = workbook.create_sheet('Тип контента')
    sheet5.append(['Тип контента', 'Процентное соотношение'])
    content_types = defaultdict(int)
    total_posts = len(parsed_data)
    for item in parsed_data:
        content_type = item['type']
        if content_type == 'image':
            content_types['image'] += 1
        elif content_type == 'video':
            content_types['video'] += 1
        elif content_type == 'text':
            content_types['text'] += 1
    for content_type, count in content_types.items():
        percentage = (count / total_posts) * 100
        sheet5.append([content_type, percentage])

    # Sheet 6: Top 3 Posts by Channel and Day (by VR Post)
    sheet6 = workbook.create_sheet('Топ-3 постов по дню')
    sheet6.append(['Дата', 'Канал', 'ID Поста', 'Ссылка', 'Категория', 'Сообщение', 'VR Post'])
    posts_by_channel_day = defaultdict(list)
    for item in parsed_data:
        date = datetime.strptime(item['date'], '%Y-%m-%d %H:%M:%S').strftime('%Y-%m-%d')
        channel = item['title']
        key = (date, channel)
        posts_by_channel_day[key].append({
            'post_id': item['post_id'],
            'link': item['link'],
            'category': item['category'],
            'message': item['message'],
            'vr_post': float(item['vr_post'])
        })
    
    for (date, channel), posts in posts_by_channel_day.items():
        top_posts = sorted(posts, key=lambda x: x['vr_post'], reverse=True)[:3]
        for post in top_posts:
            sheet6.append([date, channel, post['post_id'], post['link'], post['category'], post['message'], post['vr_post']])

    # 1) Расчет процента каждой категории для Sheet 6 (отдельная таблица)
    sheet7 = workbook.create_sheet('Процент категорий в Топ-3')
    sheet7.append(['Категория', 'Количество постов', 'Процент'])
    category_counts_top3 = defaultdict(int)
    total_top3_posts = 0
    for (date, channel), posts in posts_by_channel_day.items():
        top_posts = sorted(posts, key=lambda x: x['vr_post'], reverse=True)[:3]
        for post in top_posts:
            category = post['category']
            category_counts_top3[category] += 1
            total_top3_posts += 1
    for cat, count in category_counts_top3.items():
        percentage = (count / total_top3_posts * 100) if total_top3_posts > 0 else 0
        sheet7.append([cat, count, f"{percentage:.2f}%"])

    # 3) Динамика типов контента по каждому дню
    sheet8 = workbook.create_sheet('Динамика типов контента по дням')
    sheet8.append(['Дата', 'Тип контента', 'Количество'])
    content_by_day = defaultdict(lambda: defaultdict(int))
    for item in parsed_data:
        date = datetime.strptime(item['date'], '%Y-%m-%d %H:%M:%S').strftime('%Y-%m-%d')
        content_type = item['type'].lower()
        if content_type in ['image', 'video', 'text']:
            content_by_day[date][content_type] += 1
    for date, types in content_by_day.items():
        for content_type, count in types.items():
            sheet8.append([date, content_type, count])

    # 4) Динамика типов контента по каждой категории по каждому дню
    sheet9 = workbook.create_sheet('Динамика типов контента по категориям')
    sheet9.append(['Дата', 'Категория', 'Тип контента', 'Количество'])
    content_by_category_day = defaultdict(lambda: defaultdict(lambda: defaultdict(int)))
    for item in parsed_data:
        date = datetime.strptime(item['date'], '%Y-%m-%d %H:%M:%S').strftime('%Y-%m-%d')
        category = item.get('category', 'N/A')
        content_type = item['type'].lower()
        if content_type in ['image', 'video', 'text']:
            content_by_category_day[date][category][content_type] += 1
    for date, categories in content_by_category_day.items():
        for category, types in categories.items():
            for content_type, count in types.items():
                sheet9.append([date, category, content_type, count])

    # 5) Динамика VRpost по категориям по каждому дню
    sheet10 = workbook.create_sheet('Динамика VRpost по категориям')
    sheet10.append(['Дата', 'Категория', 'Средний VRpost'])
    vr_by_category_day = defaultdict(lambda: defaultdict(list))
    for item in parsed_data:
        date = datetime.strptime(item['date'], '%Y-%m-%d %H:%M:%S').strftime('%Y-%m-%d')
        category = item.get('category', 'N/A')
        vr_by_category_day[date][category].append(float(item['vr_post']))
    for date, categories in vr_by_category_day.items():
        for category, vrs in categories.items():
            avg_vr = mean(vrs) if vrs else 0
            sheet10.append([date, category, avg_vr])

    # 6) Динамика количества постов по каждой категории по каждому дню
    sheet11 = workbook.create_sheet('Динамика постов по категориям')
    sheet11.append(['Дата', 'Категория', 'Количество постов'])
    posts_by_category_day = defaultdict(lambda: defaultdict(int))
    for item in parsed_data:
        date = datetime.strptime(item['date'], '%Y-%m-%d %H:%M:%S').strftime('%Y-%m-%d')
        category = item.get('category', 'N/A')
        posts_by_category_day[date][category] += 1
    for date, categories in posts_by_category_day.items():
        for category, count in categories.items():
            sheet11.append([date, category, count])

    # 7) ТОП-25 публикаций по каждой категории по VRpost
    sheet12 = workbook.create_sheet('ТОП-25 постов по категориям')
    sheet12.append(['Дата', 'Канал', 'ID Поста', 'Ссылка', 'Категория', 'Сообщение', 'VR Post'])
    posts_by_category = defaultdict(list)
    for item in parsed_data:
        category = item.get('category', 'N/A')
        posts_by_category[category].append({
            'date': datetime.strptime(item['date'], '%Y-%m-%d %H:%M:%S').strftime('%Y-%m-%d'),
            'title': item['title'],
            'post_id': item['post_id'],
            'link': item['link'],
            'category': category,
            'message': item['message'],
            'vr_post': float(item['vr_post'])
        })
    for category, posts in posts_by_category.items():
        top_25_posts = sorted(posts, key=lambda x: x['vr_post'], reverse=True)[:25]
        for post in top_25_posts:
            sheet12.append([post['date'], post['title'], post['post_id'], post['link'], post['category'], post['message'], post['vr_post']])

    # Save the workbook
    current_time = datetime.now().strftime('%Y%m%d_%H%M%S')
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=telegram_data_{min_date}_to_{max_date}_{current_time}.xlsx'
    workbook.save(response)

    return response

async def get_post_details(request):
    if request.method == 'GET':
        post_id = request.GET.get('post_id')
        channel_id = request.GET.get('channel_id')
        page = int(request.GET.get('page', 1))
        limit = int(request.GET.get('limit', 10))

        if not post_id or not channel_id:
            logger.error(f"Missing or empty post_id or channel_id: post_id={post_id}, channel_id={channel_id}")
            return JsonResponse({'error': 'Missing post_id or channel_id'}, status=400)

        client = TelegramClient('session_name', api_id, api_hash)
        try:
            if not client.is_connected():
                await client.connect()
            if not await client.is_user_authorized():
                logger.error("Telegram client is not authorized. Please authorize the client first.")
                return JsonResponse({'error': 'Telegram client is not authorized. Please authorize the client first.'}, status=401)

            entity = await client.get_entity(int(channel_id))
            post = await client.get_messages(entity, ids=int(post_id))

            if not post:
                logger.error(f"Post {post_id} not found for channel {channel_id}")
                return JsonResponse({'error': 'Post not found'}, status=404)

            if isinstance(post, types.Message):
                reactions = []
                if hasattr(post, 'reactions') and post.reactions:
                    for reaction in post.reactions.results:
                        reaction_type = reaction.reaction
                        logger.debug(f"Reaction type: {type(reaction_type).__name__}, Reaction: {reaction_type}")
                        if isinstance(reaction_type, types.ReactionEmoji):
                            reactions.append({
                                'emoticon': reaction_type.emoticon,
                                'count': reaction.count
                            })
                        elif isinstance(reaction_type, types.ReactionPaid):
                            reactions.append({
                                'emoticon': '💎',
                                'count': reaction.count
                            })
                        else:
                            logger.warning(f"Unknown reaction type: {type(reaction_type).__name__}")
                            reactions.append({
                                'emoticon': '[Неизвестная реакция]',
                                'count': reaction.count
                            })

                comments_data = []
                if hasattr(post, 'replies') and post.replies and post.replies.replies > 0:
                    # Получаем все комментарии
                    all_comments = []
                    async for comment in client.iter_messages(entity, reply_to=post.id):
                        if isinstance(comment, types.Message):
                            author = await comment.get_sender()
                            author_name = author.username if author and hasattr(author, 'username') else 'Аноним'

                            comment_replies = comment.replies.replies if hasattr(comment, 'replies') and comment.replies and comment.replies.replies is not None else 0
                            comment_forwards = comment.forwards if hasattr(comment, 'forwards') and comment.forwards is not None else 0

                            all_comments.append({
                                'message': comment.message or '[Без текста]',
                                'date': comment.date.strftime('%Y-%m-%d %H:%M:%S') if comment.date else 'N/A',
                                'author': author_name,
                                'forwards': comment_forwards,
                                'replies': comment_replies
                            })
                    logger.debug(f"Fetched {len(all_comments)} comments for post {post_id}")

                    # Применяем пагинацию на стороне Python
                    start_idx = (page - 1) * limit
                    end_idx = start_idx + limit
                    paginated_comments = all_comments[start_idx:end_idx]
                    total_comments = len(all_comments)
                else:
                    logger.debug(f"No replies found for post {post_id}")
                    paginated_comments = []
                    total_comments = 0

                response_data = {
                    'reactions': reactions,
                    'comments': paginated_comments,
                    'total_comments': total_comments
                }
                logger.debug(f"Returning response: {response_data}")
                return JsonResponse(response_data)
            else:
                logger.error(f"Invalid post data for post {post_id}")
                return JsonResponse({'error': 'Invalid post data'}, status=400)
        except Exception as e:
            logger.error(f"Error in get_post_details: {str(e)}")
            return JsonResponse({'error': str(e)}, status=500)
        finally:
            if client.is_connected():
                await client.disconnect()
    return JsonResponse({'error': 'Invalid request method'}, status=405)

def update_post_category(request):
    """
    Обновляет категорию поста, но не переобучает модель.
    Переобучение будет выполнено позже, когда пользователь нажмёт "Применить изменения".
    """
    if request.method == 'POST':
        logger.info(f"Received request to update category: {request.POST}")
        data_id = request.POST.get('data_id')
        post_id = request.POST.get('post_id')
        new_category = request.POST.get('category')

        if not all([data_id, post_id, new_category]):
            logger.error("Missing required fields")
            return JsonResponse({'error': 'Missing required fields'}, status=400)

        file_path = f"temp_data/{data_id}.json"
        logger.info(f"Reading data from {file_path}")
        try:
            with open(file_path, 'r', encoding='utf-8') as f:
                parsed_data = json.load(f)
        except FileNotFoundError:
            logger.error(f"Data file not found: {file_path}")
            return JsonResponse({'error': 'Data file not found'}, status=404)
        except Exception as e:
            logger.error(f"Error reading data file: {str(e)}")
            return JsonResponse({'error': f'Error reading data file: {str(e)}'}, status=500)

        updated = False
        for item in parsed_data:
            if str(item['post_id']) == post_id:
                item['category'] = new_category
                updated = True
                break

        if not updated:
            logger.error(f"Post not found: post_id={post_id}")
            return JsonResponse({'error': 'Post not found'}, status=404)

        logger.info(f"Writing updated data to {file_path}")
        try:
            with open(file_path, 'w', encoding='utf-8') as f:
                json.dump(parsed_data, f)
        except Exception as e:
            logger.error(f"Error writing to data file: {str(e)}")
            return JsonResponse({'error': f'Error writing to data file: {str(e)}'}, status=500)

        logger.info("Category updated successfully")
        return JsonResponse({'success': True, 'message': 'Category updated. Click "Apply Changes" to retrain the model.'})
    logger.error("Invalid request method")
    return JsonResponse({'error': 'Invalid request method'}, status=405)

def apply_changes(request):
    """
    Переобучает модель на основе обновлённых данных после нажатия кнопки "Применить изменения".
    """
    global model, vectorizer
    if request.method == 'POST':
        logger.info(f"Received request to apply changes: {request.POST}")
        data_id = request.POST.get('data_id')

        if not data_id:
            logger.error("Missing data_id")
            return JsonResponse({'error': 'Missing data_id'}, status=400)

        file_path = f"temp_data/{data_id}.json"
        logger.info(f"Reading data from {file_path}")
        try:
            with open(file_path, 'r', encoding='utf-8') as f:
                parsed_data = json.load(f)
        except FileNotFoundError:
            logger.error(f"Data file not found: {file_path}")
            return JsonResponse({'error': 'Data file not found'}, status=404)
        except Exception as e:
            logger.error(f"Error reading data file: {str(e)}")
            return JsonResponse({'error': f'Error reading data file: {str(e)}'}, status=500)

        # Формируем данные для переобучения
        new_data = [{
            'Text': item['message'],
            'Category ': item['category']
        } for item in parsed_data]

        logger.info("Retraining model with updated data")
        try:
            update_model(new_data, temp_data_path=file_path)
            # Перезагружаем модель после обучения
            model, vectorizer = load_model()
            logger.info("Model retrained successfully")
            return JsonResponse({'success': True, 'message': 'Model retrained successfully'})
        except Exception as e:
            logger.error(f"Error retraining model: {str(e)}")
            return JsonResponse({'error': f'Error retraining model: {str(e)}'}, status=500)

    logger.error("Invalid request method")
    return JsonResponse({'error': 'Invalid request method'}, status=405)

def analytics_dashboard(request):
    data_id = request.GET.get('data_id')
    print("Received data_id in analytics_dashboard:", data_id)
    if not data_id:
        return HttpResponse("No data ID provided.", status=400)

    file_path = f"temp_data/{data_id}.json"
    try:
        with open(file_path, 'r', encoding='utf-8') as f:
            parsed_data = json.load(f)
        print("Successfully loaded data for data_id:", data_id)
    except FileNotFoundError:
        print("File not found for data_id:", data_id)
        return HttpResponse("Data file not found.", status=404)

    # Prepare data for calculations
    days_map = {
        'Понедельник': 'Monday',
        'Вторник': 'Tuesday',
        'Среда': 'Wednesday',
        'Четверг': 'Thursday',
        'Пятница': 'Friday',
        'Суббота': 'Saturday',
        'Воскресенье': 'Sunday'
    }
    
    # 1. Publications by Day of Week
    publications_by_day = defaultdict(int)
    for item in parsed_data:
        date = datetime.strptime(item['date'], '%Y-%m-%d %H:%M:%S')
        day_name = date.strftime('%A')
        for ru_day, en_day in days_map.items():
            if en_day == day_name:
                publications_by_day[ru_day] += 1
    publications_by_day_data = dict(publications_by_day)

    # 2. Average ER Post by Day
    er_by_day = defaultdict(list)
    for item in parsed_data:
        date = datetime.strptime(item['date'], '%Y-%m-%d %H:%M:%S')
        day_name = date.strftime('%A')
        for ru_day, en_day in days_map.items():
            if en_day == day_name:
                try:
                    er_by_day[ru_day].append(float(item.get('er_post', 0)))
                except (ValueError, TypeError):
                    er_by_day[ru_day].append(0)
    er_by_day_data = {day: mean(ers) for day, ers in er_by_day.items() if ers}

    # 3. Average VR Post by Day
    vr_by_day = defaultdict(list)
    for item in parsed_data:
        date = datetime.strptime(item['date'], '%Y-%m-%d %H:%M:%S')
        day_name = date.strftime('%A')
        for ru_day, en_day in days_map.items():
            if en_day == day_name:
                try:
                    vr_by_day[ru_day].append(float(item.get('vr_post', 0)))
                except (ValueError, TypeError):
                    vr_by_day[ru_day].append(0)
    vr_by_day_data = {day: mean(vrs) for day, vrs in vr_by_day.items() if vrs}

    # 4. Content Type Distribution
    content_types = defaultdict(int)
    total_posts = len(parsed_data)
    for item in parsed_data:
        content_type = item.get('type', 'unknown').lower()
        if content_type in ['image', 'video', 'text']:
            content_types[content_type] += 1
    content_types_data = {content_type: (count / total_posts) * 100 for content_type, count in content_types.items() if count > 0}

    # 5. Top 3 Posts by Channel and Day (by VR Post)
    top_posts_data = []
    posts_by_channel_day = defaultdict(list)
    for item in parsed_data:
        date = datetime.strptime(item['date'], '%Y-%m-%d %H:%M:%S').strftime('%Y-%m-%d')
        channel = item['title']
        key = (date, channel)
        posts_by_channel_day[key].append({
            'post_id': item['post_id'],
            'message': item['message'],
            'vr_post': float(item['vr_post'])
        })
    
    for (date, channel), posts in posts_by_channel_day.items():
        top_posts = sorted(posts, key=lambda x: x['vr_post'], reverse=True)[:3]
        for post in top_posts:
            top_posts_data.append({
                'date': date,
                'channel': channel,
                'post_id': post['post_id'],
                'message': post['message'],
                'vr_post': post['vr_post']
            })

    # Ensure all days are present
    for day in days_map.keys():
        if day not in publications_by_day_data:
            publications_by_day_data[day] = 0
        if day not in er_by_day_data:
            er_by_day_data[day] = 0
        if day not in vr_by_day_data:
            vr_by_day_data[day] = 0

    # Debug output
    print("Publications by Day:", publications_by_day_data)
    print("ER by Day:", er_by_day_data)
    print("VR by Day:", vr_by_day_data)
    print("Content Types:", content_types_data)
    print("Top Posts:", top_posts_data)

    context = {
        'data_id': data_id,
        'publications_by_day': json.dumps(publications_by_day_data),
        'er_by_day': json.dumps(er_by_day_data),
        'vr_by_day': json.dumps(vr_by_day_data),
        'content_types': json.dumps(content_types_data),
        'top_posts': top_posts_data,
    }
    print("Serialized Publications by Day:", json.dumps(publications_by_day_data))
    print("Serialized ER by Day:", json.dumps(er_by_day_data))
    print("Serialized VR by Day:", json.dumps(vr_by_day_data))
    print("Serialized Content Types:", json.dumps(content_types_data))
    # Исправляем путь к шаблону
    return render(request, 'myapp/telegram/charts.html', context)


def export_model_view(request):
    """
    Экспортирует модель и векторизатор в указанную директорию.
    """
    if request.method == 'GET':
        export_path = request.GET.get('path', 'exported_model')
        success = export_model(export_path)
        if success:
            return JsonResponse({'success': True, 'message': f'Model exported to {export_path}'})
        return JsonResponse({'success': False, 'error': 'Failed to export model'}, status=500)
    return JsonResponse({'error': 'Invalid request method'}, status=405)


@login_required
async def telegram_daily_view(request):
    filters = request.GET.getlist('filter')
    category_filters = request.GET.getlist('category_filter')
    start_date = request.GET.get('start_date', '')
    end_date = request.GET.get('end_date', '')
    sort_by = request.GET.get('sort_by', 'post_id')
    sort_direction = request.GET.get('sort_direction', 'desc')

    form = TelegramForm(initial={
        'start_date': start_date,
        'end_date': end_date,
    })

    # Асинхронный запрос к базе с предварительной загрузкой channel
    posts = await sync_to_async(lambda: TelegramPost.objects.select_related('channel').all())()

    if start_date:
        start_date = datetime.strptime(start_date, '%Y-%m-%d').replace(tzinfo=timezone.utc)
        posts = await sync_to_async(lambda: posts.filter(date__gte=start_date))()
    if end_date:
        end_date = datetime.strptime(end_date, '%Y-%m-%d').replace(hour=23, minute=59, second=59, tzinfo=timezone.utc)
        posts = await sync_to_async(lambda: posts.filter(date__lte=end_date))()
    if filters and 'all' not in filters:
        posts = await sync_to_async(lambda: posts.filter(channel__title__in=filters))()
    if category_filters and 'all' not in category_filters:
        posts = await sync_to_async(lambda: posts.filter(category__in=category_filters))()

    reverse = sort_direction == 'desc'
    if sort_by == 'channel':
        order_field = '-channel__title' if reverse else 'channel__title'
        posts = await sync_to_async(lambda: posts.order_by(order_field))()
    elif sort_by == 'postid':
        order_field = '-post_id' if reverse else 'post_id'
        posts = await sync_to_async(lambda: posts.order_by(order_field))()
    elif sort_by == 'date':
        order_field = '-date' if reverse else 'date'
        posts = await sync_to_async(lambda: posts.order_by(order_field))()
    elif sort_by == 'category':
        order_field = '-category' if reverse else 'category'
        posts = await sync_to_async(lambda: posts.order_by(order_field))()

    # Загружаем все посты в список
    posts_list = await sync_to_async(lambda: list(posts))()

    # Получение уникальных значений
    unique_titles = await sync_to_async(
        lambda: sorted(list(set(TelegramPost.objects.select_related('channel').values_list('channel__title', flat=True))))
    )()
    unique_categories_in_data = await sync_to_async(
        lambda: sorted(list(set(TelegramPost.objects.values_list('category', flat=True).exclude(category__isnull=True).exclude(category='N/A'))))
    )()

    # Пагинация
    paginator = Paginator(posts_list, 20)  # Используем список, а не QuerySet
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    # Аналитика
    days_map = {
        'Понедельник': 'Monday',
        'Вторник': 'Tuesday',
        'Среда': 'Wednesday',
        'Четверг': 'Thursday',
        'Пятница': 'Friday',
        'Суббота': 'Saturday',
        'Воскресенье': 'Sunday'
    }
    publications_by_day = defaultdict(int)
    er_by_day = defaultdict(list)
    vr_by_day = defaultdict(list)
    content_types = defaultdict(int)
    top_posts_data = []

    for post in posts_list:
        date = post.date
        day_name = date.strftime('%A')
        for ru_day, en_day in days_map.items():
            if en_day == day_name:
                publications_by_day[ru_day] += 1
                er_by_day[ru_day].append(float(post.er_post or 0))
                vr_by_day[ru_day].append(float(post.vr_post or 0))
        content_type = post.post_type.lower()
        if content_type in ['image', 'video', 'text']:
            content_types[content_type] += 1

    er_by_day_data = {day: mean(ers) for day, ers in er_by_day.items() if ers}
    vr_by_day_data = {day: mean(vrs) for day, vrs in vr_by_day.items() if vrs}
    content_types_data = {content_type: (count / len(posts_list)) * 100 for content_type, count in content_types.items() if count > 0}

    posts_by_channel_day = defaultdict(list)
    for post in posts_list:
        date = post.date.strftime('%Y-%m-%d')
        channel = post.channel.title
        key = (date, channel)
        posts_by_channel_day[key].append({
            'post_id': post.post_id,
            'message': post.message or 'N/A',
            'vr_post': float(post.vr_post or 0),
            'channel': channel  # Сохраняем channel для использования в top_posts
        })

    for (date, channel), posts in posts_by_channel_day.items():
        top_posts = sorted(posts, key=lambda x: x['vr_post'], reverse=True)[:3]
        for post in top_posts:
            top_posts_data.append({
                'date': date,
                'channel': post['channel'],
                'post_id': post['post_id'],
                'message': post['message'],
                'vr_post': post['vr_post']
            })

    for day in days_map.keys():
        if day not in publications_by_day:
            publications_by_day[day] = 0
        if day not in er_by_day_data:
            er_by_day_data[day] = 0
        if day not in vr_by_day_data:
            vr_by_day_data[day] = 0

    context = {
        'form': form,
        'parsed_data': page_obj,
        'filters': filters,
        'category_filters': category_filters,
        'unique_titles': unique_titles,
        'start_date': start_date,
        'end_date': end_date,
        'sort_by': sort_by,
        'sort_direction': sort_direction,
        'unique_categories': unique_categories_in_data,
        'unique_categories_in_data': unique_categories_in_data,
        'publications_by_day': json.dumps(dict(publications_by_day)),
        'er_by_day': json.dumps(er_by_day_data),
        'vr_by_day': json.dumps(vr_by_day_data),
        'content_types': json.dumps(content_types_data),
        'top_posts': top_posts_data,
    }
    return await sync_to_async(render)(request, 'myapp/telegram/telegram_daily.html', context)