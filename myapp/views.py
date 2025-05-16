# Django imports
from django.shortcuts import render
from django.http import HttpResponse
from django.http import JsonResponse
from django.core.paginator import Paginator
from asgiref.sync import sync_to_async

# Django users imports
from django.contrib.auth.decorators import login_required
from django.contrib.auth.decorators import user_passes_test
from django.contrib.auth import logout
from django.shortcuts import redirect
from django.contrib.auth import authenticate, login
from django.contrib import messages
from django.contrib.auth.views import PasswordChangeView

# Forms
from .forms import TelegramForm, CustomAuthenticationForm, CustomPasswordChangeForm
# Regular expressions
import re

# Date and time
from datetime import datetime
from datetime import timezone
import time

# Telegram related imports
from telethon import TelegramClient
from telethon import utils
from telethon import types
from telethon import functions

# Logging
import logging

# Excel processing
import openpyxl

# JSON handling
import json

# UUID generation
import uuid

# Operating system utilities
import os

# Environment variables
from dotenv import load_dotenv

# Machine learning imports
from machine_learning.train_model import load_model
from machine_learning.train_model import predict_category
from machine_learning.train_model import update_model
from machine_learning.train_model import get_unique_categories
from machine_learning.train_model import export_model

# Image processing
from PIL import Image
import io
from collections import defaultdict
from statistics import mean

from asgiref.sync import sync_to_async

load_dotenv()
api_id = os.getenv("API_ID")  
api_hash = os.getenv("API_HASH")  

# –ù–∞—Å—Ç—Ä–æ–π–∫–∞ –ª–æ–≥–∏—Ä–æ–≤–∞–Ω–∏—è
logger = logging.getLogger(__name__)

# –ó–∞–≥—Ä—É–∑–∫–∞ –º–æ–¥–µ–ª–∏ –∏ –≤–µ–∫—Ç–æ—Ä–∏–∑–∞—Ç–æ—Ä–∞ –ø—Ä–∏ —Å—Ç–∞—Ä—Ç–µ
model, vectorizer = load_model()

def index(request):
    return render(request, 'myapp/base.html')

def index(request):
    return render(request, 'myapp/base.html')

def custom_logout(request):
    logout(request)
    messages.success(request, "–í—ã —É—Å–ø–µ—à–Ω–æ –≤—ã—à–ª–∏!")
    return redirect('index')

def custom_login(request):
    if request.method == 'POST':
        form = CustomAuthenticationForm(request, data=request.POST)
        if form.is_valid():
            username = form.cleaned_data['username']
            password = form.cleaned_data['password']
            user = authenticate(request, username=username, password=password)
            if user is not None:
                login(request, user)
                next_url = request.POST.get('next', request.GET.get('next', 'index'))
                messages.success(request, "–í—ã —É—Å–ø–µ—à–Ω–æ –≤–æ—à–ª–∏!")
                return redirect(next_url)
            else:
                messages.error(request, "–ù–µ–≤–µ—Ä–Ω–æ–µ –∏–º—è –ø–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª—è –∏–ª–∏ –ø–∞—Ä–æ–ª—å.")
        else:
            messages.error(request, "–û—à–∏–±–∫–∞ –≤ –¥–∞–Ω–Ω—ã—Ö —Ñ–æ—Ä–º—ã.")
    else:
        form = CustomAuthenticationForm()
    return render(request, 'myapp/login.html', {'form': form})

def password_change_done(request):
    return render(request, 'myapp/password_change_done.html')

class CustomPasswordChangeView(PasswordChangeView):
    form_class = CustomPasswordChangeForm
    template_name = 'myapp/password_change.html'
    success_url = '/password_change_done/'  # –û—Å—Ç–∞–≤–ª—è–µ–º –∫–∞–∫ –µ—Å—Ç—å, –Ω–æ —Ç–µ–ø–µ—Ä—å –º–∞—Ä—à—Ä—É—Ç –±—É–¥–µ—Ç —Ä–∞–±–æ—Ç–∞—Ç—å

    def form_valid(self, form):
        response = super().form_valid(form)
        messages.success(self.request, "–í–∞—à –ø–∞—Ä–æ–ª—å —É—Å–ø–µ—à–Ω–æ –∏–∑–º–µ–Ω—ë–Ω!")
        return response

    def form_invalid(self, form):
        messages.error(self.request, "–û—à–∏–±–∫–∞ –ø—Ä–∏ —Å–º–µ–Ω–µ –ø–∞—Ä–æ–ª—è. –ü—Ä–æ–≤–µ—Ä—å—Ç–µ –≤–≤–µ–¥—ë–Ω–Ω—ã–µ –¥–∞–Ω–Ω—ã–µ.")
        return self.render_to_response(self.get_context_data(form=form))

def cleanup_temp_data(folder='temp_data', max_age_seconds=86400):
    """
    –£–¥–∞–ª—è–µ—Ç —Ñ–∞–π–ª—ã –≤ —É–∫–∞–∑–∞–Ω–Ω–æ–π –ø–∞–ø–∫–µ, –∫–æ—Ç–æ—Ä—ã–µ —Å—Ç–∞—Ä—à–µ max_age_seconds.
    –ü–æ —É–º–æ–ª—á–∞–Ω–∏—é —É–¥–∞–ª—è–µ—Ç —Ñ–∞–π–ª—ã —Å—Ç–∞—Ä—à–µ –æ–¥–Ω–æ–≥–æ –¥–Ω—è (86400 —Å–µ–∫—É–Ω–¥).
    """
    now = time.time()
    for filename in os.listdir(folder):
        file_path = os.path.join(folder, filename)
        if os.path.isfile(file_path):
            file_age = now - os.path.getmtime(file_path)
            if file_age > max_age_seconds:
                os.remove(file_path)
                logger.debug(f"Deleted old file: {file_path}")

def extract_username(url):
    match = re.search(r'https://t.me/([^/?]+)', url)
    if match:
        return match.group(1)
    raise ValueError("Invalid Telegram URL")


async def fetch_telegram_data(channel_url, start_date=None, end_date=None):
    global model, vectorizer
    client = TelegramClient('session_name', api_id, api_hash)
    try:
        if not client.is_connected():
            await client.connect()
        if not await client.is_user_authorized():
            logger.warning("–ö–ª–∏–µ–Ω—Ç Telegram –Ω–µ –∞–≤—Ç–æ—Ä–∏–∑–æ–≤–∞–Ω. –ü–æ–∂–∞–ª—É–π—Å—Ç–∞, –∞–≤—Ç–æ—Ä–∏–∑—É–π—Ç–µ—Å—å –≤—Ä—É—á–Ω—É—é.")
            print("–í–≤–µ–¥–∏—Ç–µ –Ω–æ–º–µ—Ä —Ç–µ–ª–µ—Ñ–æ–Ω–∞ (–Ω–∞–ø—Ä–∏–º–µ—Ä, +79991234567): ")
            phone = input().strip()
            await client.start(phone=phone)

        if start_date:
            start_date = datetime.combine(start_date, datetime.min.time()).replace(tzinfo=timezone.utc)
            logger.debug(f"–î–∞—Ç–∞ –Ω–∞—á–∞–ª–∞ —É—Å—Ç–∞–Ω–æ–≤–ª–µ–Ω–∞ –Ω–∞: {start_date}")
        if end_date:
            end_date = datetime.combine(end_date, datetime.max.time()).replace(tzinfo=timezone.utc)
            logger.debug(f"–î–∞—Ç–∞ –æ–∫–æ–Ω—á–∞–Ω–∏—è —É—Å—Ç–∞–Ω–æ–≤–ª–µ–Ω–∞ –Ω–∞: {end_date}")

        if "https://t.me/" in channel_url:
            username = channel_url.split('/')[-1]
        elif "https://web.telegram.org/k/#" in channel_url:
            channel_id = int(channel_url.split('#')[-1])
            username = utils.get_peer_id(types.PeerChannel(channel_id))
        else:
            username = extract_username(channel_url)

        logger.debug(f"–ü–æ–ª—É—á–µ–Ω–∏–µ —Å—É—â–Ω–æ—Å—Ç–∏ –¥–ª—è username: {username}")
        entity = await client.get_entity(username)
        data = []
        combined_message = None

        # –ü–æ–ª—É—á–µ–Ω–∏–µ –∫–æ–ª–∏—á–µ—Å—Ç–≤–∞ –ø–æ–¥–ø–∏—Å—á–∏–∫–æ–≤
        subscriber_count = 0
        try:
            full_channel = await client(functions.channels.GetFullChannelRequest(channel=entity))
            subscriber_count = int(getattr(full_channel.full_chat, 'participants_count', 0))
            logger.debug(f"–ö–æ–ª–∏—á–µ—Å—Ç–≤–æ –ø–æ–¥–ø–∏—Å—á–∏–∫–æ–≤ –∏–∑ GetFullChannelRequest –¥–ª—è {entity.title}: {subscriber_count}")
        except Exception as e:
            logger.error(f"–û—à–∏–±–∫–∞ –ø–æ–ª—É—á–µ–Ω–∏—è –∫–æ–ª–∏—á–µ—Å—Ç–≤–∞ –ø–æ–¥–ø–∏—Å—á–∏–∫–æ–≤ –¥–ª—è –∫–∞–Ω–∞–ª–∞ {entity.id}: {str(e)}")
            subscriber_count = 0

        if subscriber_count == 0:
            logger.warning(f"–ù–µ —É–¥–∞–ª–æ—Å—å –ø–æ–ª—É—á–∏—Ç—å –≤–∞–ª–∏–¥–Ω–æ–µ –∫–æ–ª–∏—á–µ—Å—Ç–≤–æ –ø–æ–¥–ø–∏—Å—á–∏–∫–æ–≤ –¥–ª—è –∫–∞–Ω–∞–ª–∞ {entity.title}. –ú–µ—Ç—Ä–∏–∫–∏ –±—É–¥—É—Ç —Ä–∞–≤–Ω—ã 0. –£–±–µ–¥–∏—Ç–µ—Å—å, —á—Ç–æ –∫–ª–∏–µ–Ω—Ç –∏–º–µ–µ—Ç —Ä–∞–∑—Ä–µ—à–µ–Ω–∏–µ '–ü—Ä–æ—Å–º–æ—Ç—Ä —Å—Ç–∞—Ç–∏—Å—Ç–∏–∫–∏ –∫–∞–Ω–∞–ª–∞'.")

        # –ó–∞–≥—Ä—É–∑–∫–∞ –∞–≤–∞—Ç–∞—Ä–∞ –∫–∞–Ω–∞–ª–∞
        avatar_path = None
        avatar_dir = os.path.join('temp_data', 'avatars')
        os.makedirs(avatar_dir, exist_ok=True)
        avatar_filename = f"channel_{entity.id}.jpg"
        avatar_full_path = os.path.join(avatar_dir, avatar_filename)

        if hasattr(entity, 'photo') and entity.photo:
            logger.debug(f"–ó–∞–≥—Ä—É–∑–∫–∞ —Ñ–æ—Ç–æ –ø—Ä–æ—Ñ–∏–ª—è –¥–ª—è –∫–∞–Ω–∞–ª–∞ {entity.id}")
            try:
                photo = await client.download_profile_photo(entity, file=avatar_full_path)
                if photo and os.path.exists(avatar_full_path):
                    logger.debug(f"–§–æ—Ç–æ –ø—Ä–æ—Ñ–∏–ª—è –∑–∞–≥—Ä—É–∂–µ–Ω–æ –≤ {avatar_full_path}")
                    try:
                        with Image.open(avatar_full_path) as img:
                            img = img.resize((50, 50), Image.LANCZOS)
                            img.save(avatar_full_path, 'JPEG', quality=85)
                        avatar_path = f"avatars/{avatar_filename}"
                        logger.info(f"–ê–≤–∞—Ç–∞—Ä –æ–±—Ä–∞–±–æ—Ç–∞–Ω –∏ —Å–æ—Ö—Ä–∞–Ω—ë–Ω –¥–ª—è –∫–∞–Ω–∞–ª–∞ {entity.id}: {avatar_path}")
                    except Exception as e:
                        logger.error(f"–û—à–∏–±–∫–∞ –æ–±—Ä–∞–±–æ—Ç–∫–∏ –∞–≤–∞—Ç–∞—Ä–∞ –¥–ª—è –∫–∞–Ω–∞–ª–∞ {entity.id}: {str(e)}")
                        avatar_path = None
                else:
                    logger.warning(f"–ê–≤–∞—Ç–∞—Ä –Ω–µ –∑–∞–≥—Ä—É–∂–µ–Ω –¥–ª—è –∫–∞–Ω–∞–ª–∞ {entity.id} (photo={photo})")
                    avatar_path = None
            except Exception as e:
                logger.error(f"–û—à–∏–±–∫–∞ –∑–∞–≥—Ä—É–∑–∫–∏ –∞–≤–∞—Ç–∞—Ä–∞ –¥–ª—è –∫–∞–Ω–∞–ª–∞ {entity.id}: {str(e)}")
                avatar_path = None
        else:
            logger.debug(f"–§–æ—Ç–æ –ø—Ä–æ—Ñ–∏–ª—è –Ω–µ–¥–æ—Å—Ç—É–ø–Ω–æ –¥–ª—è –∫–∞–Ω–∞–ª–∞ {entity.id}")
            avatar_path = None

        total_engagement = 0
        total_comments = 0
        post_count = 0

        async for post in client.iter_messages(entity, reverse=True, offset_date=start_date):
            if end_date and post.date and post.date > end_date:
                logger.debug(f"–ü—Ä–æ–ø—É—Å–∫ –ø–æ—Å—Ç–∞ {post.id} —Ç–∞–∫ –∫–∞–∫ –¥–∞—Ç–∞ {post.date} –ø–æ–∑–∂–µ end_date {end_date}")
                break

            # –û–ø—Ä–µ–¥–µ–ª–µ–Ω–∏–µ —Ç–∏–ø–∞ –ø–æ—Å—Ç–∞
            post_type = 'text'
            if post.media:
                if isinstance(post.media, types.MessageMediaPhoto):
                    post_type = 'image'
                elif isinstance(post.media, types.MessageMediaDocument) and post.media.document.mime_type.startswith('video'):
                    post_type = 'video'

            message_text = post.message if post.message else 'N/A'
            # –ò–≥–Ω–æ—Ä–∏—Ä—É–µ–º –ø–æ—Å—Ç—ã, –∫–æ—Ç–æ—Ä—ã–µ —è–≤–ª—è—é—Ç—Å—è —Ç–æ–ª—å–∫–æ –º–µ–¥–∏–∞ –±–µ–∑ —Ç–µ–∫—Å—Ç–∞ –∏–ª–∏ –≤–ª–æ–∂–µ–Ω–Ω—ã–º–∏ –º–µ–¥–∏–∞ –≤ –≥—Ä—É–ø–ø–µ
            if message_text == 'N/A' and (not post.grouped_id or (combined_message and combined_message.get('post_id') != post.id)):
                logger.debug(f"–ò–≥–Ω–æ—Ä–∏—Ä–æ–≤–∞–Ω–∏–µ –ø–æ—Å—Ç–∞ {post.id}: —Ç–æ–ª—å–∫–æ –º–µ–¥–∏–∞ –±–µ–∑ —Ç–µ–∫—Å—Ç–∞ –∏–ª–∏ –≤–ª–æ–∂–µ–Ω–Ω—ã–π —Ñ–∞–π–ª (—Ç–∏–ø: {post_type})")
                if post.grouped_id and combined_message:
                    combined_message['message'] += f"\n[{post_type.capitalize()}: {post.id}]"
                    combined_message['link'] = f"https://t.me/{entity.username}/{post.id}"
                continue

            if post.media and isinstance(post.media, types.MessageMediaPhoto) and not post.message and combined_message and combined_message.get('message') != 'N/A':
                combined_message['message'] += f"\n[–ò–∑–æ–±—Ä–∞–∂–µ–Ω–∏–µ: {post.media.photo.id}]"
                combined_message['link'] = f"https://t.me/{entity.username}/{post.id}"
                continue

            comments_count = post.replies.replies if hasattr(post, 'replies') and post.replies else 0
            reactions_count = sum(reaction.count for reaction in post.reactions.results) if hasattr(post, 'reactions') and post.reactions else 0
            forwards_count = post.forwards if hasattr(post, 'forwards') else 0
            views = post.views if hasattr(post, 'views') else 0

            logger.debug(f"–ü–æ—Å—Ç {post.id}: comments={comments_count}, reactions={reactions_count}, forwards={forwards_count}, views={views}, type={post_type}")

            category = predict_category(message_text, model, vectorizer) if message_text != 'N/A' else 'N/A'

            # –†–∞—Å—á—ë—Ç –º–µ—Ç—Ä–∏–∫
            post_engagement = reactions_count + forwards_count + comments_count
            total_engagement += post_engagement
            total_comments += comments_count
            post_count += 1

            er_post = (post_engagement / subscriber_count * 100) if subscriber_count > 0 and post_engagement > 0 else 0
            er_view = (post_engagement / views * 100) if views > 0 and post_engagement > 0 else 0
            vr_post = (views / subscriber_count * 100) if subscriber_count > 0 and views > 0 else 0
            logger.debug(f"–ü–æ—Å—Ç {post.id}: engagement={post_engagement}, er_post={er_post}%, er_view={er_view}%, vr_post={vr_post}%")

            combined_message = {
                'title': entity.title,
                'id': entity.id,
                'post_id': post.id,
                'date': post.date.strftime('%Y-%m-%d %H:%M:%S') if post.date else 'N/A',
                'message': message_text,
                'link': f"https://t.me/{entity.username}/{post.id}" if entity.username else f"https://t.me/c/{entity.id}/{post.id}",
                'views': views,
                'reactions': reactions_count,
                'forwards': forwards_count,
                'comments_count': comments_count,
                'category': category,
                'avatar': avatar_path,
                'er_post': round(er_post, 2),
                'er_view': round(er_view, 2),
                'vr_post': round(vr_post, 2),
                'type': post_type,
                'subscribers': subscriber_count
            }

            data.append(combined_message)

        # –†–∞—Å—á—ë—Ç TR
        tr = (total_comments / subscriber_count / post_count * 100) if post_count > 0 and subscriber_count > 0 else 0
        logger.debug(f"–†–∞—Å—á—ë—Ç TR: total_comments={total_comments}, subscriber_count={subscriber_count}, post_count={post_count}, tr={tr}%")
        for item in data:
            item['tr'] = round(tr, 2)

        data_id = str(uuid.uuid4())
        os.makedirs('temp_data', exist_ok=True)
        with open(f'temp_data/{data_id}.json', 'w', encoding='utf-8') as f:
            json.dump(data, f)

        logger.debug(f"–ü–æ–ª—É—á–µ–Ω—ã –¥–∞–Ω–Ω—ã–µ –¥–ª—è {len(data)} –ø–æ—Å—Ç–æ–≤ –∏–∑ –∫–∞–Ω–∞–ª–∞ {entity.title}")
        logger.debug(f"ID –¥–∞–Ω–Ω—ã—Ö: {data_id}")
        logger.debug(f"–ü—É—Ç–∏ –∞–≤–∞—Ç–∞—Ä–æ–≤ –≤ –¥–∞–Ω–Ω—ã—Ö: {[item['avatar'] for item in data if item['avatar']]}")

        return data, data_id
    except Exception as e:
        logger.error(f"–û—à–∏–±–∫–∞ –ø–æ–ª—É—á–µ–Ω–∏—è –¥–∞–Ω–Ω—ã—Ö –¥–ª—è –∫–∞–Ω–∞–ª–∞ {channel_url}: {str(e)}")
        raise
    finally:
        if client.is_connected():
            await client.disconnect()

# –í–Ω—É—Ç—Ä–∏ telegram_view

# ... (–æ—Å—Ç–∞–ª—å–Ω—ã–µ –∏–º–ø–æ—Ä—Ç—ã –æ—Å—Ç–∞—é—Ç—Å—è –±–µ–∑ –∏–∑–º–µ–Ω–µ–Ω–∏–π)

@login_required
async def telegram_view(request):
    cleanup_temp_data()

    filters = request.GET.getlist('filter')
    category_filters = request.GET.getlist('category_filter')
    channel_url = request.GET.get('channel_url', '')
    start_date = request.GET.get('start_date', '')
    end_date = request.GET.get('end_date', '')
    data_id = request.GET.get('data_id', '')
    sort_by = request.GET.get('sort_by', 'date')
    sort_direction = request.GET.get('sort_direction', 'desc')

    if request.method == 'POST':
        form = TelegramForm(request.POST)
        if form.is_valid():
            channel_urls = form.cleaned_data['channel_url'].strip().split('\n')
            start_date = form.cleaned_data['start_date']
            end_date = form.cleaned_data['end_date']
            all_data = []
            data_id_list = []

            for channel_url in channel_urls:
                channel_url = channel_url.strip()
                if channel_url:
                    data, data_id = await fetch_telegram_data(channel_url, start_date, end_date)
                    all_data.extend(data)
                    data_id_list.append(data_id)
            
            combined_data_id = str(uuid.uuid4())
            with open(f'temp_data/{combined_data_id}.json', 'w', encoding='utf-8') as f:
                json.dump(all_data, f)

            unique_titles = sorted(list(set(item['title'] for item in all_data)))
            logger.debug(f"Combined data passed to template: {len(all_data)} posts")
            logger.debug(f"Combined Data ID passed to template: {combined_data_id}")
            
            filtered_data = [
                item for item in all_data 
                if (not filters or 'all' in filters or item['title'] in filters) and
                   (not category_filters or 'all' in category_filters or item['category'] in category_filters)
            ]
            
            reverse = sort_direction == 'desc'
            if sort_by == 'channel':
                filtered_data.sort(key=lambda x: x['title'], reverse=reverse)
            elif sort_by == 'postid':
                filtered_data.sort(key=lambda x: x['post_id'], reverse=reverse)
            elif sort_by == 'date':
                filtered_data.sort(key=lambda x: x['date'], reverse=reverse)
            elif sort_by == 'category':
                filtered_data.sort(key=lambda x: x['category'], reverse=reverse)

            unique_categories_in_data = sorted(list(set(item['category'] for item in filtered_data if item['category'])))
            paginator = Paginator(filtered_data, 20)
            page_number = request.GET.get('page')
            page_obj = paginator.get_page(page_number)

            # –ü–æ–¥–≥–æ—Ç–æ–≤–∫–∞ –¥–∞–Ω–Ω—ã—Ö –¥–ª—è –≥—Ä–∞—Ñ–∏–∫–æ–≤
            days_map = {
                '–ü–æ–Ω–µ–¥–µ–ª—å–Ω–∏–∫': 'Monday',
                '–í—Ç–æ—Ä–Ω–∏–∫': 'Tuesday',
                '–°—Ä–µ–¥–∞': 'Wednesday',
                '–ß–µ—Ç–≤–µ—Ä–≥': 'Thursday',
                '–ü—è—Ç–Ω–∏—Ü–∞': 'Friday',
                '–°—É–±–±–æ—Ç–∞': 'Saturday',
                '–í–æ—Å–∫—Ä–µ—Å–µ–Ω—å–µ': 'Sunday'
            }
            publications_by_day = defaultdict(int)
            er_by_day = defaultdict(list)
            vr_by_day = defaultdict(list)
            content_types = defaultdict(int)
            top_posts_data = []

            for item in all_data:
                date = datetime.strptime(item['date'], '%Y-%m-%d %H:%M:%S')
                day_name = date.strftime('%A')
                for ru_day, en_day in days_map.items():
                    if en_day == day_name:
                        publications_by_day[ru_day] += 1
                        try:
                            er_by_day[ru_day].append(float(item.get('er_post', 0)))
                        except (ValueError, TypeError):
                            er_by_day[ru_day].append(0)
                        try:
                            vr_by_day[ru_day].append(float(item.get('vr_post', 0)))
                        except (ValueError, TypeError):
                            vr_by_day[ru_day].append(0)
                content_type = item.get('type', 'unknown').lower()
                if content_type in ['image', 'video', 'text']:
                    content_types[content_type] += 1

            er_by_day_data = {day: mean(ers) for day, ers in er_by_day.items() if ers}
            vr_by_day_data = {day: mean(vrs) for day, vrs in vr_by_day.items() if vrs}
            content_types_data = {content_type: (count / len(all_data)) * 100 for content_type, count in content_types.items() if count > 0}

            posts_by_channel_day = defaultdict(list)
            for item in all_data:
                date = datetime.strptime(item['date'], '%Y-%m-%d %H:%M:%S').strftime('%Y-%m-%d')
                channel = item['title']
                key = (date, channel)
                posts_by_channel_day[key].append({
                    'post_id': item['post_id'],
                    'message': item['message'],
                    'vr_post': float(item['vr_post'])
                })
            
            for (date, channel), posts in posts_by_channel_day.items():
                top_posts = sorted(posts, key=lambda x: x['vr_post'], reverse=True)[:3]
                for post in top_posts:
                    top_posts_data.append({
                        'date': date,
                        'channel': channel,
                        'post_id': post['post_id'],
                        'message': post['message'],
                        'vr_post': post['vr_post']
                    })

            for day in days_map.keys():
                if day not in publications_by_day:
                    publications_by_day[day] = 0
                if day not in er_by_day_data:
                    er_by_day_data[day] = 0
                if day not in vr_by_day_data:
                    vr_by_day_data[day] = 0

            context = {
                'data_id': data_id,
                'form': form,
                'parsed_data': page_obj,
                'data_id': combined_data_id,
                'filters': filters,
                'category_filters': category_filters,
                'unique_titles': unique_titles,
                'channel_url': request.POST.get('channel_url'),
                'start_date': request.POST.get('start_date'),
                'end_date': request.POST.get('end_date'),
                'sort_by': sort_by,
                'sort_direction': sort_direction,
                'unique_categories': get_unique_categories(),
                'unique_categories_in_data': unique_categories_in_data,
                'publications_by_day': json.dumps(dict(publications_by_day)),
                'er_by_day': json.dumps(er_by_day_data),
                'vr_by_day': json.dumps(vr_by_day_data),
                'content_types': json.dumps(content_types_data),
                'top_posts': top_posts_data,
            }
            print("Rendering telegram_main.html with data_id:", data_id)  # –û—Ç–ª–∞–¥–∫–∞
            # –û–±–æ—Ä–∞—á–∏–≤–∞–µ–º render –≤ sync_to_async
            return await sync_to_async(render)(request, 'myapp/telegram/telegram_main.html', context)
    else:
        form = TelegramForm(initial={
            'channel_url': channel_url,
            'start_date': start_date,
            'end_date': end_date,
        })
    
    parsed_data = []
    if data_id:
        try:
            with open(f'temp_data/{data_id}.json', 'r', encoding='utf-8') as f:
                parsed_data = json.load(f)
            logger.debug(f"Loaded parsed data with {len(parsed_data)} posts")
        except FileNotFoundError:
            logger.error(f"Data file not found: temp_data/{data_id}.json")
    
    unique_titles = sorted(list(set(item['title'] for item in parsed_data)))
    
    filtered_data = [
        item for item in parsed_data 
        if (not filters or 'all' in filters or item['title'] in filters) and
           (not category_filters or 'all' in category_filters or item['category'] in category_filters)
    ]
    
    reverse = sort_direction == 'desc'
    if sort_by == 'channel':
        filtered_data.sort(key=lambda x: x['title'], reverse=reverse)
    elif sort_by == 'postid':
        filtered_data.sort(key=lambda x: x['post_id'], reverse=reverse)
    elif sort_by == 'date':
        filtered_data.sort(key=lambda x: x['date'], reverse=reverse)
    elif sort_by == 'category':
        filtered_data.sort(key=lambda x: x['category'], reverse=reverse)

    unique_categories_in_data = sorted(list(set(item['category'] for item in filtered_data if item['category'])))
    paginator = Paginator(filtered_data, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    # –ü–æ–¥–≥–æ—Ç–æ–≤–∫–∞ –¥–∞–Ω–Ω—ã—Ö –¥–ª—è –≥—Ä–∞—Ñ–∏–∫–æ–≤
    days_map = {
        '–ü–æ–Ω–µ–¥–µ–ª—å–Ω–∏–∫': 'Monday',
        '–í—Ç–æ—Ä–Ω–∏–∫': 'Tuesday',
        '–°—Ä–µ–¥–∞': 'Wednesday',
        '–ß–µ—Ç–≤–µ—Ä–≥': 'Thursday',
        '–ü—è—Ç–Ω–∏—Ü–∞': 'Friday',
        '–°—É–±–±–æ—Ç–∞': 'Saturday',
        '–í–æ—Å–∫—Ä–µ—Å–µ–Ω—å–µ': 'Sunday'
    }
    publications_by_day = defaultdict(int)
    er_by_day = defaultdict(list)
    vr_by_day = defaultdict(list)
    content_types = defaultdict(int)
    top_posts_data = []

    for item in parsed_data:
        date = datetime.strptime(item['date'], '%Y-%m-%d %H:%M:%S')
        day_name = date.strftime('%A')
        for ru_day, en_day in days_map.items():
            if en_day == day_name:
                publications_by_day[ru_day] += 1
                try:
                    er_by_day[ru_day].append(float(item.get('er_post', 0)))
                except (ValueError, TypeError):
                    er_by_day[ru_day].append(0)
                try:
                    vr_by_day[ru_day].append(float(item.get('vr_post', 0)))
                except (ValueError, TypeError):
                    vr_by_day[ru_day].append(0)
        content_type = item.get('type', 'unknown').lower()
        if content_type in ['image', 'video', 'text']:
            content_types[content_type] += 1

    er_by_day_data = {day: mean(ers) for day, ers in er_by_day.items() if ers}
    vr_by_day_data = {day: mean(vrs) for day, vrs in vr_by_day.items() if vrs}
    content_types_data = {content_type: (count / len(parsed_data)) * 100 for content_type, count in content_types.items() if count > 0}

    posts_by_channel_day = defaultdict(list)
    for item in parsed_data:
        date = datetime.strptime(item['date'], '%Y-%m-%d %H:%M:%S').strftime('%Y-%m-%d')
        channel = item['title']
        key = (date, channel)
        posts_by_channel_day[key].append({
            'post_id': item['post_id'],
            'message': item['message'],
            'vr_post': float(item['vr_post'])
        })
    
    for (date, channel), posts in posts_by_channel_day.items():
        top_posts = sorted(posts, key=lambda x: x['vr_post'], reverse=True)[:3]
        for post in top_posts:
            top_posts_data.append({
                'date': date,
                'channel': channel,
                'post_id': post['post_id'],
                'message': post['message'],
                'vr_post': post['vr_post']
            })

    for day in days_map.keys():
        if day not in publications_by_day:
            publications_by_day[day] = 0
        if day not in er_by_day_data:
            er_by_day_data[day] = 0
        if day not in vr_by_day_data:
            vr_by_day_data[day] = 0

    context = {
        'form': form,
        'parsed_data': page_obj,
        'data_id': data_id,
        'filters': filters,
        'category_filters': category_filters,
        'unique_titles': unique_titles,
        'channel_url': channel_url,
        'start_date': start_date,
        'end_date': end_date,
        'sort_by': sort_by,
        'sort_direction': sort_direction,
        'unique_categories': get_unique_categories(),
        'unique_categories_in_data': unique_categories_in_data,
        'publications_by_day': json.dumps(dict(publications_by_day)),
        'er_by_day': json.dumps(er_by_day_data),
        'vr_by_day': json.dumps(vr_by_day_data),
        'content_types': json.dumps(content_types_data),
        'top_posts': top_posts_data,
    }
    # –û–±–æ—Ä–∞—á–∏–≤–∞–µ–º render –≤ sync_to_async
    return await sync_to_async(render)(request, 'myapp/telegram/telegram_main.html', context)


def export_to_excel(request):
    data_id = request.GET.get('data_id')
    if not data_id:
        return HttpResponse("No data ID provided.", status=400)

    file_path = f"temp_data/{data_id}.json"
    try:
        with open(file_path, 'r', encoding='utf-8') as f:
            parsed_data = json.load(f)
    except FileNotFoundError:
        return HttpResponse("Data file not found.", status=404)

    workbook = openpyxl.Workbook()
    
    # Get date range for filename
    dates = [datetime.strptime(item['date'], '%Y-%m-%d %H:%M:%S') for item in parsed_data]
    min_date = min(dates).strftime('%Y%m%d')
    max_date = max(dates).strftime('%Y%m%d')

    # Sheet 1: Original Data
    sheet1 = workbook.active
    sheet1.title = 'Original Data'
    headers = ['–ö–∞–Ω–∞–ª', 'ID –ö–∞–Ω–∞–ª–∞', 'ID –ü–æ—Å—Ç–∞', '–î–∞—Ç–∞', '–°–æ–æ–±—â–µ–Ω–∏–µ', '–ö–∞—Ç–µ–≥–æ—Ä–∏—è', '–°—Å—ã–ª–∫–∞', '–ü—Ä–æ—Å–º–æ—Ç—Ä–æ–≤', '–†–µ–∞–∫—Ü–∏–∏', '–ü–µ—Ä–µ—Å—ã–ª–∫–∏', '–ö–æ–º–º–µ–Ω—Ç–∞—Ä–∏–∏']
    sheet1.append(headers)

    for item in parsed_data:
        row = [
            item['title'], 
            item['id'], 
            item['post_id'], 
            item['date'], 
            item['message'], 
            item['category'],
            item['link'], 
            item['views'], 
            item['reactions'], 
            item['forwards'],
            item['comments_count']
        ]
        sheet1.append(row)

    # Prepare data for calculations
    days_map = {
        '–ü–æ–Ω–µ–¥–µ–ª—å–Ω–∏–∫': 'Monday',
        '–í—Ç–æ—Ä–Ω–∏–∫': 'Tuesday',
        '–°—Ä–µ–¥–∞': 'Wednesday',
        '–ß–µ—Ç–≤–µ—Ä–≥': 'Thursday',
        '–ü—è—Ç–Ω–∏—Ü–∞': 'Friday',
        '–°—É–±–±–æ—Ç–∞': 'Saturday',
        '–í–æ—Å–∫—Ä–µ—Å–µ–Ω—å–µ': 'Sunday'
    }
    
    # Sheet 2: Publications by Day of Week
    sheet2 = workbook.create_sheet('Publications by Day')
    sheet2.append(['–î–µ–Ω—å –Ω–µ–¥–µ–ª–∏', '–ö–æ–ª–∏—á–µ—Å—Ç–≤–æ –ø—É–±–ª–∏–∫–∞—Ü–∏–π'])
    publications_by_day = defaultdict(int)
    for item in parsed_data:
        date = datetime.strptime(item['date'], '%Y-%m-%d %H:%M:%S')
        day_name = date.strftime('%A')
        # Convert English day name to Russian
        for ru_day, en_day in days_map.items():
            if en_day == day_name:
                publications_by_day[ru_day] += 1
    for day, count in publications_by_day.items():
        sheet2.append([day, count])

    # Sheet 3: Average ER Post by Day
    sheet3 = workbook.create_sheet('Average ER Post by Day')
    sheet3.append(['–î–µ–Ω—å –Ω–µ–¥–µ–ª–∏', '–°—Ä–µ–¥–Ω–∏–π ER Post'])
    er_by_day = defaultdict(list)
    for item in parsed_data:
        date = datetime.strptime(item['date'], '%Y-%m-%d %H:%M:%S')
        day_name = date.strftime('%A')
        for ru_day, en_day in days_map.items():
            if en_day == day_name:
                er_by_day[ru_day].append(float(item['er_post']))
    for day, ers in er_by_day.items():
        sheet3.append([day, mean(ers)])

    # Sheet 4: Average VR Post by Day
    sheet4 = workbook.create_sheet('Average VR Post by Day')
    sheet4.append(['–î–µ–Ω—å –Ω–µ–¥–µ–ª–∏', '–°—Ä–µ–¥–Ω–∏–π VR Post'])
    vr_by_day = defaultdict(list)
    for item in parsed_data:
        date = datetime.strptime(item['date'], '%Y-%m-%d %H:%M:%S')
        day_name = date.strftime('%A')
        for ru_day, en_day in days_map.items():
            if en_day == day_name:
                vr_by_day[ru_day].append(float(item['vr_post']))
    for day, vrs in vr_by_day.items():
        sheet4.append([day, mean(vrs)])

    # Sheet 5: Content Type Distribution
    sheet5 = workbook.create_sheet('Content Type Distribution')
    sheet5.append(['–¢–∏–ø –∫–æ–Ω—Ç–µ–Ω—Ç–∞', '–ü—Ä–æ—Ü–µ–Ω—Ç–Ω–æ–µ —Å–æ–æ—Ç–Ω–æ—à–µ–Ω–∏–µ'])
    content_types = defaultdict(int)
    total_posts = len(parsed_data)
    for item in parsed_data:
        content_type = item['type']
        if content_type == 'image':
            content_types['image'] += 1
        elif content_type == 'video':
            content_types['video'] += 1
        elif content_type == 'text':
            content_types['text'] += 1
    for content_type, count in content_types.items():
        percentage = (count / total_posts) * 100
        sheet5.append([content_type, percentage])

    # Sheet 6: Top 3 Posts by Channel and Day (by VR Post)
    sheet6 = workbook.create_sheet('Top 3 Posts by Channel-Day')
    sheet6.append(['–î–∞—Ç–∞', '–ö–∞–Ω–∞–ª', 'ID –ü–æ—Å—Ç–∞', '–°–æ–æ–±—â–µ–Ω–∏–µ', 'VR Post'])
    posts_by_channel_day = defaultdict(list)
    for item in parsed_data:
        date = datetime.strptime(item['date'], '%Y-%m-%d %H:%M:%S').strftime('%Y-%m-%d')
        channel = item['title']
        key = (date, channel)
        posts_by_channel_day[key].append({
            'post_id': item['post_id'],
            'message': item['message'],
            'vr_post': float(item['vr_post'])
        })
    
    for (date, channel), posts in posts_by_channel_day.items():
        # Sort by VR Post and take top 3
        top_posts = sorted(posts, key=lambda x: x['vr_post'], reverse=True)[:3]
        for post in top_posts:
            sheet6.append([date, channel, post['post_id'], post['message'], post['vr_post']])

    # Save the workbook
    current_time = datetime.now().strftime('%Y%m%d_%H%M%S')
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=telegram_data_{min_date}_to_{max_date}_{current_time}.xlsx'
    workbook.save(response)

    return response

async def get_post_details(request):
    if request.method == 'GET':
        post_id = request.GET.get('post_id')
        channel_id = request.GET.get('channel_id')
        page = int(request.GET.get('page', 1))
        limit = int(request.GET.get('limit', 10))

        if not post_id or not channel_id:
            logger.error(f"Missing or empty post_id or channel_id: post_id={post_id}, channel_id={channel_id}")
            return JsonResponse({'error': 'Missing post_id or channel_id'}, status=400)

        client = TelegramClient('session_name', api_id, api_hash)
        try:
            if not client.is_connected():
                await client.connect()
            if not await client.is_user_authorized():
                logger.error("Telegram client is not authorized. Please authorize the client first.")
                return JsonResponse({'error': 'Telegram client is not authorized. Please authorize the client first.'}, status=401)

            entity = await client.get_entity(int(channel_id))
            post = await client.get_messages(entity, ids=int(post_id))

            if not post:
                logger.error(f"Post {post_id} not found for channel {channel_id}")
                return JsonResponse({'error': 'Post not found'}, status=404)

            if isinstance(post, types.Message):
                reactions = []
                if hasattr(post, 'reactions') and post.reactions:
                    for reaction in post.reactions.results:
                        reaction_type = reaction.reaction
                        logger.debug(f"Reaction type: {type(reaction_type).__name__}, Reaction: {reaction_type}")
                        if isinstance(reaction_type, types.ReactionEmoji):
                            reactions.append({
                                'emoticon': reaction_type.emoticon,
                                'count': reaction.count
                            })
                        elif isinstance(reaction_type, types.ReactionPaid):
                            reactions.append({
                                'emoticon': 'üíé',
                                'count': reaction.count
                            })
                        else:
                            logger.warning(f"Unknown reaction type: {type(reaction_type).__name__}")
                            reactions.append({
                                'emoticon': '[–ù–µ–∏–∑–≤–µ—Å—Ç–Ω–∞—è —Ä–µ–∞–∫—Ü–∏—è]',
                                'count': reaction.count
                            })

                comments_data = []
                if hasattr(post, 'replies') and post.replies and post.replies.replies > 0:
                    # –ü–æ–ª—É—á–∞–µ–º –≤—Å–µ –∫–æ–º–º–µ–Ω—Ç–∞—Ä–∏–∏
                    all_comments = []
                    async for comment in client.iter_messages(entity, reply_to=post.id):
                        if isinstance(comment, types.Message):
                            author = await comment.get_sender()
                            author_name = author.username if author and hasattr(author, 'username') else '–ê–Ω–æ–Ω–∏–º'

                            comment_replies = comment.replies.replies if hasattr(comment, 'replies') and comment.replies and comment.replies.replies is not None else 0
                            comment_forwards = comment.forwards if hasattr(comment, 'forwards') and comment.forwards is not None else 0

                            all_comments.append({
                                'message': comment.message or '[–ë–µ–∑ —Ç–µ–∫—Å—Ç–∞]',
                                'date': comment.date.strftime('%Y-%m-%d %H:%M:%S') if comment.date else 'N/A',
                                'author': author_name,
                                'forwards': comment_forwards,
                                'replies': comment_replies
                            })
                    logger.debug(f"Fetched {len(all_comments)} comments for post {post_id}")

                    # –ü—Ä–∏–º–µ–Ω—è–µ–º –ø–∞–≥–∏–Ω–∞—Ü–∏—é –Ω–∞ —Å—Ç–æ—Ä–æ–Ω–µ Python
                    start_idx = (page - 1) * limit
                    end_idx = start_idx + limit
                    paginated_comments = all_comments[start_idx:end_idx]
                    total_comments = len(all_comments)
                else:
                    logger.debug(f"No replies found for post {post_id}")
                    paginated_comments = []
                    total_comments = 0

                response_data = {
                    'reactions': reactions,
                    'comments': paginated_comments,
                    'total_comments': total_comments
                }
                logger.debug(f"Returning response: {response_data}")
                return JsonResponse(response_data)
            else:
                logger.error(f"Invalid post data for post {post_id}")
                return JsonResponse({'error': 'Invalid post data'}, status=400)
        except Exception as e:
            logger.error(f"Error in get_post_details: {str(e)}")
            return JsonResponse({'error': str(e)}, status=500)
        finally:
            if client.is_connected():
                await client.disconnect()
    return JsonResponse({'error': 'Invalid request method'}, status=405)

def update_post_category(request):
    """
    –û–±–Ω–æ–≤–ª—è–µ—Ç –∫–∞—Ç–µ–≥–æ—Ä–∏—é –ø–æ—Å—Ç–∞, –Ω–æ –Ω–µ –ø–µ—Ä–µ–æ–±—É—á–∞–µ—Ç –º–æ–¥–µ–ª—å.
    –ü–µ—Ä–µ–æ–±—É—á–µ–Ω–∏–µ –±—É–¥–µ—Ç –≤—ã–ø–æ–ª–Ω–µ–Ω–æ –ø–æ–∑–∂–µ, –∫–æ–≥–¥–∞ –ø–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª—å –Ω–∞–∂–º—ë—Ç "–ü—Ä–∏–º–µ–Ω–∏—Ç—å –∏–∑–º–µ–Ω–µ–Ω–∏—è".
    """
    if request.method == 'POST':
        logger.info(f"Received request to update category: {request.POST}")
        data_id = request.POST.get('data_id')
        post_id = request.POST.get('post_id')
        new_category = request.POST.get('category')

        if not all([data_id, post_id, new_category]):
            logger.error("Missing required fields")
            return JsonResponse({'error': 'Missing required fields'}, status=400)

        file_path = f"temp_data/{data_id}.json"
        logger.info(f"Reading data from {file_path}")
        try:
            with open(file_path, 'r', encoding='utf-8') as f:
                parsed_data = json.load(f)
        except FileNotFoundError:
            logger.error(f"Data file not found: {file_path}")
            return JsonResponse({'error': 'Data file not found'}, status=404)
        except Exception as e:
            logger.error(f"Error reading data file: {str(e)}")
            return JsonResponse({'error': f'Error reading data file: {str(e)}'}, status=500)

        updated = False
        for item in parsed_data:
            if str(item['post_id']) == post_id:
                item['category'] = new_category
                updated = True
                break

        if not updated:
            logger.error(f"Post not found: post_id={post_id}")
            return JsonResponse({'error': 'Post not found'}, status=404)

        logger.info(f"Writing updated data to {file_path}")
        try:
            with open(file_path, 'w', encoding='utf-8') as f:
                json.dump(parsed_data, f)
        except Exception as e:
            logger.error(f"Error writing to data file: {str(e)}")
            return JsonResponse({'error': f'Error writing to data file: {str(e)}'}, status=500)

        logger.info("Category updated successfully")
        return JsonResponse({'success': True, 'message': 'Category updated. Click "Apply Changes" to retrain the model.'})
    logger.error("Invalid request method")
    return JsonResponse({'error': 'Invalid request method'}, status=405)

def apply_changes(request):
    """
    –ü–µ—Ä–µ–æ–±—É—á–∞–µ—Ç –º–æ–¥–µ–ª—å –Ω–∞ –æ—Å–Ω–æ–≤–µ –æ–±–Ω–æ–≤–ª—ë–Ω–Ω—ã—Ö –¥–∞–Ω–Ω—ã—Ö –ø–æ—Å–ª–µ –Ω–∞–∂–∞—Ç–∏—è –∫–Ω–æ–ø–∫–∏ "–ü—Ä–∏–º–µ–Ω–∏—Ç—å –∏–∑–º–µ–Ω–µ–Ω–∏—è".
    """
    global model, vectorizer
    if request.method == 'POST':
        logger.info(f"Received request to apply changes: {request.POST}")
        data_id = request.POST.get('data_id')

        if not data_id:
            logger.error("Missing data_id")
            return JsonResponse({'error': 'Missing data_id'}, status=400)

        file_path = f"temp_data/{data_id}.json"
        logger.info(f"Reading data from {file_path}")
        try:
            with open(file_path, 'r', encoding='utf-8') as f:
                parsed_data = json.load(f)
        except FileNotFoundError:
            logger.error(f"Data file not found: {file_path}")
            return JsonResponse({'error': 'Data file not found'}, status=404)
        except Exception as e:
            logger.error(f"Error reading data file: {str(e)}")
            return JsonResponse({'error': f'Error reading data file: {str(e)}'}, status=500)

        # –§–æ—Ä–º–∏—Ä—É–µ–º –¥–∞–Ω–Ω—ã–µ –¥–ª—è –ø–µ—Ä–µ–æ–±—É—á–µ–Ω–∏—è
        new_data = [{
            'Text': item['message'],
            'Category ': item['category']
        } for item in parsed_data]

        logger.info("Retraining model with updated data")
        try:
            update_model(new_data, temp_data_path=file_path)
            # –ü–µ—Ä–µ–∑–∞–≥—Ä—É–∂–∞–µ–º –º–æ–¥–µ–ª—å –ø–æ—Å–ª–µ –æ–±—É—á–µ–Ω–∏—è
            model, vectorizer = load_model()
            logger.info("Model retrained successfully")
            return JsonResponse({'success': True, 'message': 'Model retrained successfully'})
        except Exception as e:
            logger.error(f"Error retraining model: {str(e)}")
            return JsonResponse({'error': f'Error retraining model: {str(e)}'}, status=500)

    logger.error("Invalid request method")
    return JsonResponse({'error': 'Invalid request method'}, status=405)

def analytics_dashboard(request):
    data_id = request.GET.get('data_id')
    print("Received data_id in analytics_dashboard:", data_id)
    if not data_id:
        return HttpResponse("No data ID provided.", status=400)

    file_path = f"temp_data/{data_id}.json"
    try:
        with open(file_path, 'r', encoding='utf-8') as f:
            parsed_data = json.load(f)
        print("Successfully loaded data for data_id:", data_id)
    except FileNotFoundError:
        print("File not found for data_id:", data_id)
        return HttpResponse("Data file not found.", status=404)

    # Prepare data for calculations
    days_map = {
        '–ü–æ–Ω–µ–¥–µ–ª—å–Ω–∏–∫': 'Monday',
        '–í—Ç–æ—Ä–Ω–∏–∫': 'Tuesday',
        '–°—Ä–µ–¥–∞': 'Wednesday',
        '–ß–µ—Ç–≤–µ—Ä–≥': 'Thursday',
        '–ü—è—Ç–Ω–∏—Ü–∞': 'Friday',
        '–°—É–±–±–æ—Ç–∞': 'Saturday',
        '–í–æ—Å–∫—Ä–µ—Å–µ–Ω—å–µ': 'Sunday'
    }
    
    # 1. Publications by Day of Week
    publications_by_day = defaultdict(int)
    for item in parsed_data:
        date = datetime.strptime(item['date'], '%Y-%m-%d %H:%M:%S')
        day_name = date.strftime('%A')
        for ru_day, en_day in days_map.items():
            if en_day == day_name:
                publications_by_day[ru_day] += 1
    publications_by_day_data = dict(publications_by_day)

    # 2. Average ER Post by Day
    er_by_day = defaultdict(list)
    for item in parsed_data:
        date = datetime.strptime(item['date'], '%Y-%m-%d %H:%M:%S')
        day_name = date.strftime('%A')
        for ru_day, en_day in days_map.items():
            if en_day == day_name:
                try:
                    er_by_day[ru_day].append(float(item.get('er_post', 0)))
                except (ValueError, TypeError):
                    er_by_day[ru_day].append(0)
    er_by_day_data = {day: mean(ers) for day, ers in er_by_day.items() if ers}

    # 3. Average VR Post by Day
    vr_by_day = defaultdict(list)
    for item in parsed_data:
        date = datetime.strptime(item['date'], '%Y-%m-%d %H:%M:%S')
        day_name = date.strftime('%A')
        for ru_day, en_day in days_map.items():
            if en_day == day_name:
                try:
                    vr_by_day[ru_day].append(float(item.get('vr_post', 0)))
                except (ValueError, TypeError):
                    vr_by_day[ru_day].append(0)
    vr_by_day_data = {day: mean(vrs) for day, vrs in vr_by_day.items() if vrs}

    # 4. Content Type Distribution
    content_types = defaultdict(int)
    total_posts = len(parsed_data)
    for item in parsed_data:
        content_type = item.get('type', 'unknown').lower()
        if content_type in ['image', 'video', 'text']:
            content_types[content_type] += 1
    content_types_data = {content_type: (count / total_posts) * 100 for content_type, count in content_types.items() if count > 0}

    # 5. Top 3 Posts by Channel and Day (by VR Post)
    top_posts_data = []
    posts_by_channel_day = defaultdict(list)
    for item in parsed_data:
        date = datetime.strptime(item['date'], '%Y-%m-%d %H:%M:%S').strftime('%Y-%m-%d')
        channel = item['title']
        key = (date, channel)
        posts_by_channel_day[key].append({
            'post_id': item['post_id'],
            'message': item['message'],
            'vr_post': float(item['vr_post'])
        })
    
    for (date, channel), posts in posts_by_channel_day.items():
        top_posts = sorted(posts, key=lambda x: x['vr_post'], reverse=True)[:3]
        for post in top_posts:
            top_posts_data.append({
                'date': date,
                'channel': channel,
                'post_id': post['post_id'],
                'message': post['message'],
                'vr_post': post['vr_post']
            })

    # Ensure all days are present
    for day in days_map.keys():
        if day not in publications_by_day_data:
            publications_by_day_data[day] = 0
        if day not in er_by_day_data:
            er_by_day_data[day] = 0
        if day not in vr_by_day_data:
            vr_by_day_data[day] = 0

    # Debug output
    print("Publications by Day:", publications_by_day_data)
    print("ER by Day:", er_by_day_data)
    print("VR by Day:", vr_by_day_data)
    print("Content Types:", content_types_data)
    print("Top Posts:", top_posts_data)

    context = {
        'data_id': data_id,
        'publications_by_day': json.dumps(publications_by_day_data),
        'er_by_day': json.dumps(er_by_day_data),
        'vr_by_day': json.dumps(vr_by_day_data),
        'content_types': json.dumps(content_types_data),
        'top_posts': top_posts_data,
    }
    print("Serialized Publications by Day:", json.dumps(publications_by_day_data))
    print("Serialized ER by Day:", json.dumps(er_by_day_data))
    print("Serialized VR by Day:", json.dumps(vr_by_day_data))
    print("Serialized Content Types:", json.dumps(content_types_data))
    # –ò—Å–ø—Ä–∞–≤–ª—è–µ–º –ø—É—Ç—å –∫ —à–∞–±–ª–æ–Ω—É
    return render(request, 'myapp/telegram/charts.html', context)


def export_model_view(request):
    """
    –≠–∫—Å–ø–æ—Ä—Ç–∏—Ä—É–µ—Ç –º–æ–¥–µ–ª—å –∏ –≤–µ–∫—Ç–æ—Ä–∏–∑–∞—Ç–æ—Ä –≤ —É–∫–∞–∑–∞–Ω–Ω—É—é –¥–∏—Ä–µ–∫—Ç–æ—Ä–∏—é.
    """
    if request.method == 'GET':
        export_path = request.GET.get('path', 'exported_model')
        success = export_model(export_path)
        if success:
            return JsonResponse({'success': True, 'message': f'Model exported to {export_path}'})
        return JsonResponse({'success': False, 'error': 'Failed to export model'}, status=500)
    return JsonResponse({'error': 'Invalid request method'}, status=405)